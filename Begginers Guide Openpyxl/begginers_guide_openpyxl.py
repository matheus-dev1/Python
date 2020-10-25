#! python3

#begginers_guide_openpyxl.py

#O programa e um guia basico do modulo openpyxl

#TODO:

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

#Vamos importar as funcoes get_column_letter e column_index_from_string no modo que nao precisamos
#o usar os prefixos para utilizar a funcao.
#imports importantes.

workbook = openpyxl.load_workbook('Book1.xlsx')
#Abrir o Workbook, ou seja, o Documento Excel com todas as planilhas que voce colocou dentro deste documento.
#Obs: A planilha nao pode possuir senha.

print(type(workbook))
#Exibir a classe desta variavel. Objeto Workbook

workbook_names = workbook.get_sheet_names()
#Obtendo todos os nomes das planilhas dentro do Workbook "Book1.xlsx"

print(workbook_names)
#Exibir os nomes das planilhas em forma de lista.

print(type(workbook_names))
#E uma lista de valores normal.

Sheet2 = workbook.get_sheet_by_name('Sheet2')
#Aqui nos buscamos por uma planilha especifica, dando o seu nome como argumento em forma de string.

print(Sheet2)

print(type(Sheet2))
#Exibindo o tipo de 'dado' esta armazenado nesta varivel. Neste caso e o <class 'openpyxl.worksheet.worksheet.Worksheet'> que significa
#que esta classe e uma PLANILHa e NAO um WORKBOOK.

print(Sheet2.title)
#Com o metodo .title nos pegamos apenas o nome desta planilha.

active_sheet = workbook.active
#O metodo active nao pode receber nenhuma argumento entao nao se usa parentes depois de .active

#Aqui nos iremos pegar a planilha ativa ou seja, a planilha que está sendo visualizada no momento 
#(ou a última visualizada antes de fechar o Excel) se chama planilha ativa (active sheet).

#Curiosidade para ver o que uma funcao faz ou como ela funciona, clique nela com o botao esquerdo
#depois com o botao direito e arraste ate o nome "peek" ira abrir uma aba, depois bvoce deve 
#arrasta o mouse ate "Peek Definition" e assim ela ira abrir o arquivo desta funcao/metodo e 
#voce vera o seu funcionamento!

print("Planilha Ativa: " + str(active_sheet))
#Exibe a planilha ativa. Atraves do metodo .active

for sheet in workbook_names:
    print(sheet)
    #Exibindo os nomes das planilhas dentro do Workbook "Book1.xlsx"

Sheet1 = workbook.get_sheet_by_name('Sheet1')
#Aqui nos buscamos uma planilha especifica dentro do workbook "book1".

print(Sheet1['A2'])
#Aqui nos iremos exibir que isso e genuinamente uma celula.

print(Sheet1['A1'].value)
#Exibir o conteudo da celula 'A1' da planilha "Sheet1" do Workbook "Book1"

print(type(Sheet1['A1'].value))
#Aqui nos iremos exibir o tipo do valor retornado pela celula. Neste caso o tipo do dado desta celula e
#um objeto DateTime

B1 = Sheet1['B1']
#Atribuicao da Celula a variavel B1, ou seja, pode usar qualquer metodo de openpyxl referente a Cell.

print('Linha: ' + str(B1.row) + ' | Coluna: ' + str(B1.column) + ' | Valor: ' + str(B1.value))
#O metodo row pode ser usado em um objeto Cell e tras um valor inteiro da linha aonde a celular esta.
#Por exemplo esta celula esta na linha 1.
#O metodo column pode ser usado em um objeto Cell e tras um valor inteiro da coluna aonde a celula esta.
#Por exemplo esta celular esta na coluna 2.
#O metodo value pode ser usado em um objeto CEll e tras um vaor do tipo referente ao dado que esta na celula.
#Por exemplo esta celula contem o valor 'Apples' entao sera retonrado o valor 'Apples' do tipo String.
#Row = Linha - Column = Coluna - Coordinate = Cordenada

print('A celula: ' + str(B1.coordinate) + ' | Valor: ' + str(B1.value))
#O metodo coordinate pode ser usado em um objeto Cell e retorna a posicao deste celula com tipo do valor string.
#Exemplo B1.coordinate = B1

print(Sheet1.cell(row = 2, column = 2))
#O metodo Cell pode ser usado em um objeto Worksheet (Planilha). Devem ser usados dois argumentos nomeados: row (linha) que tem que ser um
#valor do tipo inteiro e column (coluna) que tem que ser um valor do tipo inteiro e no final retorna um objeto Cell com as determinadas
#posicoes passadas como argumento. Neste caso ele retornara B2

print(Sheet1.cell(row = 2, column = 2).value)
#Aqui nos iremos usar o metodo CEll para obter a celular e em seguinda com o valor de retorno de Cell() usaremos o metodo 
#.value na celula B2.

print("Teste da funcao zip()")
for Row, Column in zip(range(1, 6 + 1), range(10, 5, -1)):
    #A funcao zip() e usada para fazer o looping de two ou mais variaveis simultaneamente.
    print(Row, Column)

for Row in range(1, 8):
    for Column in range(1, 4):
        print(Sheet1.cell(row = Row, column = Column).value, end = ' ')
        print(Sheet1.cell(row = Row, column = Column).coordinate, end = ' ')
        print("(" + get_column_letter(Column) + ")", end =' ')
        #Aqui nos iremos exibir todas as 21 celular preenchidas por algum dado.
    print()

print(Sheet1.max_row)
#Exibe o numero da maior Linha da planilha. Nao possui parenteses para argumentos.
#Obs: O metodo get_highest_row() foi descontinuado.

print(Sheet1.max_column)
#Exibe o numero da maior Coluna da planilha. Nao possui parenteses para argumentos.
#O metodo get_highest_column() foi descontinuado.

print("Coluna a partir de um numero: " + get_column_letter(2))
print("Coluna a partir de um numero: " + get_column_letter(3))
#A funcao get_column_letter nos retornamos uma letra da coluna referente ao numero passado
#Exemplo: Se eu passar como argumento o numero 1, ele ira me retornar o letra "A" porque nos 
#pegamos a primeira coluna que no caso e a coluna A.
#Obs: Não é necessário ter um workbook carregado para usar essas funções.

print("Ultimo letra da maior coluna: " + get_column_letter(Sheet1.max_column))
#Aqui nos iremos encontrar a maior coluna da Planilha "Sheet1" e passar como argumento o seu retorno
#a funcao get_column_letter assim ele vai nos retornar em qual coluna o maior numero esta.

print("Coluna a partir de uma Letra: " + str(column_index_from_string("C")))
#A funcao column_index_from_string vai retonar uma numero a partir de uma letra.
#Neste caso eu passei o a letra C que e a terceira coluna de uma planilha entao me sera retornado
#o numero 3.

for Line in range(1, 8):
    for Column in range(1, 4):
        print("Coluna: " + get_column_letter(Column), end = ' ')
        print("Linha: " + str(column_index_from_string(get_column_letter(Line))))
        #Com a juncao da funcao column_index_from_string + get_column_letter(Line) + um laco de repeticao
        #for a gente consegue obter todas as colunas e todas as linhas
    print()

print(tuple(Sheet1["A1":"C6"]))#Objeto Generator
#Slice - Vou ate a Celula A1 ate a Celula C6, passando por todos durante este caminho.

for lines in Sheet1["A1": "D6"]:
    print(lines)#A cada lopping Sheet2 vei me passar as linhas inteiras de A1 ate D1, depois A2, ate D2
    #depois A3 ate D3, depois A4 ate D4, depois A5 ate D5 e por fim A6 ate D6.
    for celula in lines:#AQui nos vamos passar cada celula para a variavel "celula".
        print(celula.coordinate, celula.value)#Entao vamos usar os metodos .coordinate no objeto cell para
        #obter a coordenada da celula e o metodo .value no objeto cell para obter o valor dentro da celula.
    print("----End of the line----")

print(Sheet1["A"])
#Obs: Na atualizacao 2.4.0 ouve uma mudanca entao, o metodo .colunns[] nao e mais usado, no contrario
#nos usamo apenas a planilhas e entre colchetes no colocamos a letra da coluna, entao ele vai me exibir
#todas as colunas preenchidas desta coluna!

print(Sheet1["1"])
#O mesmo caso ocorre para as linhas!!!

for cell in Sheet1["A"]:
    print(cell.value)
    #Exibimos todos os valores da coluna 'A'.