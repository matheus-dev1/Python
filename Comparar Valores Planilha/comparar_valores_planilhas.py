#! python3

#comparar_valores_planilhas.py

#O programa deve compara as colunas e tabelas de duas planilhas e ver quais celulas de cada planilha possui o maior valor e exibir apenas os
#os maiores valores.

#TODO: 1. Abrir os dois arquivo Excel
#TODO: 2. Abrir as dusa planilhas que voce quer comparar
#TODO: 3. Encontrar os dados de cada planilha
#TODO: 4. Comparar os dados de cada celula de cada planilha e mostrar qual celula o maior valor de cada celula e qual a planilha.

import openpyxl
from openpyxl.descriptors.base import DateTime

print('Opening Workbook1...')

workbook = openpyxl.load_workbook('dadosExemplo.xlsx')
#Abrir o Arquivo Excel.

mainSheet = workbook.get_sheet_by_name("Dados")
#Abrir a Planilha de uso.

print('Reading Columns and Rows - Workbook...')

#______________________________________________

print('Opening Workbook2...')

workbook2 = openpyxl.load_workbook('dadosExemplo2.xlsx')
#Abrir o Arquivo Excel.

mainSheet2 = workbook2.get_sheet_by_name("Dados2")
#Abrir a Planilha de uso.

print('Reading Columns and Rows - Workbook2...')

Sheets = [mainSheet, mainSheet2]

dadosTotal = []

for UniSheet in Sheets:
    mainSheetData = []
    dadosTotal.append(mainSheetData)
    for row in range(1, UniSheet.max_row + 1):
        #[[[],[],[],[]],[[],[],[]]]
        #[0][0][0]
        dadosRow = []

        dadosRow.append(UniSheet["A" + str(row)].value)
        dadosRow.append(UniSheet["B" + str(row)].value)
        dadosRow.append(UniSheet["C" + str(row)].value)

        mainSheetData.append(dadosRow)

print(dadosTotal)

for row in range(len(dadosTotal[0])):
    for column in range(len(dadosTotal[0])):
        if dadosTotal[0][row][column] > dadosTotal[1][row][column]:
            print(str(dadosTotal[0][row][column]) + ' - ' + str(mainSheet.cell(row = row + 1, column = column + 1).coordinate))
        else:
            print(str(dadosTotal[1][row][column]) + ' - ' + str(mainSheet2.cell(row = row + 1, column = column + 1).coordinate))