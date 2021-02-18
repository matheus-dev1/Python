#!python3

#ExcelMultiplicacao.py

#O programa deve multiplicar fazer uma multiplicaxao de n x n 
#definidos pelo usuario e armazenar em um planila.

#TODO:

import openpyxl, sys, os
from openpyxl.styles import Font

workbook = openpyxl.Workbook()

sheet = workbook['Sheet']

multiplicador = sys.argv[1]

print(multiplicador)

for row in range(1, int(multiplicador) + 1):
    for column in range(1, int(multiplicador) + 1):
        if column == 1 or row == 1:
            sheet.cell(row = row, column = column).font = Font(bold=True)
            sheet.cell(row = row, column = column).value = (row * column)
        elif column == 1 and row == 1:
            sheet.cell(row = row, column = column).value = None
        else:
            sheet.cell(row = row, column = column).value = (row * column)

workbook.save('ExcelMultiplicacao' + multiplicador + '.xlsx')

os.startfile('ExcelMultiplicacao' + multiplicador + '.xlsx')