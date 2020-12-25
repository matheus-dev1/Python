#! python3

#CriandoEscrevendo2.py

#Neste programa eu crio um Workbook vazio com um planilha(que e pedrao e vem com o nome Sheet), e vou
#alterar e salvar a plailha 

#TODO:

import openpyxl, os

os.chdir('C:\Matheus\Study\IT\Python\Arquivos Python\Criando e escrevendo em planilhas\Criando e escrevendo 2')
#Altero o local do arquivo para que o Workbook seja criado no local onde este arquivo esta.

workbook = openpyxl.Workbook()
#Cria um Workbook vazio.

workbook_names = workbook.sheetnames
#O metodo sheetnames mostra todos os nomes de todas as plnilhas deste workbook

print(workbook_names)
#Exibe os nomes das plnilhas deste Workbook

sheet = workbook.active
#Aqui eu seleciona o Workbook da planilha e passo o metodo active que vai mostar a planilha ativa que no caso
#e Sheet porque e a unica!

print(sheet.title)
#Regatando o titulo da planilha.

sheet.title = "Sheet Alterada"
#Altero a propriedade Title que eh o titulo da Planilha

print(sheet.title)
#Exibo a alterecao.

workbook.save('Escrevendo2.xlsx')
#SAlvamos o arquivo propriamente dito.