#! python3

#Criar_Remover_Planilhas.py

#O programa descreve como usar os metodos create_Sheet e remove_sheet do modulo openpyxl.

#TODO:

import openpyxl, os

os.chdir('C:\Matheus\Study\IT\Python\Arquivos Python\Criando e escrevendo em planilhas\Criar e Remover Planilhas')

workbook = openpyxl.Workbook()
#Criacao de um Workbook vazio.

sheet_names = workbook.sheetnames
#Busca o nome de todas as planilhas dentro deste Workbook

print(sheet_names)
#Exibe o nome de todas as planilha deste workbook

workbook.create_sheet()
#Pegamos o nosso Workbook e usamos o metodo create_sheet() para criar uma planilha neste workbook.
#Obs: Caso voce nao coloque o parametro title= entao ele ira renomear como Sheet, porem caso tenha 
#uma planilha com o nome Sheet ele vai adicionar um numero 1, e assim subsequentemente.

workbook.create_sheet(index=0, title='Primeira Planilha Criada')
#Argumento nomeado.
#O index ou indice vai ser a posicao da nossa planilha no workbook, neste caso eu coloquei como primeira
#posicao.
#O title ou titulo vai nomear a planilha em que eu quero criar.

workbook.create_sheet(index=2, title="Planilha no indice 2")

workbook.remove_sheet(workbook['Primeira Planilha Criada'])
#Para retornar uma Worksheet ou seja uma planilha atraves do nome, nos usamos o Workbook, e entre
#chave o nome da planilha
#O metodo remove_sheet remove uma planlha com o argumento worksheet pasado como parametro.

workbook.remove_sheet(workbook['Sheet1'])

sheet_names = workbook.sheetnames

print(sheet_names)

workbook.save('Criar_Remover_Planilhas.xlsx')