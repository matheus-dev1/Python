#! python3

#Web, Text and Clipboard.py

#O programa deve poder abrir sites, arquivos de texto ou pegar texto do clipboard e passar para o
#seu excel, e salvar.

#TODO:

import openpyxl, os, sys, requests, bs4, pyperclip

os.chdir('C:\Matheus\Study\IT\Python\Arquivos Python\Criando e escrevendo em planilhas\Sites, arquivos de texto e clipboard')

workbook = openpyxl.Workbook()

sheet = workbook['Sheet']

choice = int(input('''Digite um numero para fazer determinado evento!
[1] Ler dados da Web.
[2] Ler dados de Arquivos de Texto.
[3] Ler dados do Clipboard.
[0] Sair.
Digite aqui a sua decisao: '''))

if choice == 1:
    if len(sys.argv) < 2:
        print('Esta faltando o site. Tente novamente.')
    elif len(sys.argv) == 2:
        url = sys.argv[1]
        object_response_requests_get = requests.get(url)
        print('Downloading the page... %s' %url)
        try:
            object_response_requests_get.raise_for_status()
            object_response_soup = bs4.BeautifulSoup(object_response_requests_get.text, 'html.parser')
            object_response_soup_Lista = str(object_response_soup).split()
            object_response_soup_len = len(object_response_soup_Lista)
            for row in range(0, object_response_soup_len):
                sheet.cell(row = row + 1, column = 1).value = object_response_soup_Lista[row]
                
            workbook.save('PlanilhaWorlbook.xlsx')
        except Exception as error:
            print(error)

elif choice == 2:
    Texto = open('texto.txt')
    Texto = Texto.read()
    TextoLista = Texto.split()
    TextoLinhas = len(Texto.split())

    for row in range(0, TextoLinhas):
        sheet.cell(row = row + 1, column = 1).value = TextoLista[row]

    workbook.save('Planilha.xlsx')

elif choice == 3:
    Texto = pyperclip.paste()
    TextoLista = Texto.split()
    TextoLinhas = len(Texto.split())

    for row in range(0, TextoLinhas):
        sheet.cell(row = row + 1, column = 1).value = TextoLista[row]
    
    workbook.save('PlanilhaClipboard.xlsx')

elif choice == 0:
    SystemExit()
