#! python3

#enviarDadosPlanilha.py

#O programa deve abrir a planilha com valores e adicionar a outra planilha vazia de forma aleatoria.

#TODO:

import openpyxl, os, string, random

os.chdir('C:\Matheus\Study\IT\Python\Arquivos Python\Criando e escrevendo em planilhas\Pegar \
dados de uma planilha e enviar para outra')

workbook = openpyxl.load_workbook('produceSales.xlsx')

sheet = workbook['Sheet']

###################################################################################################

newWorkbook = openpyxl.Workbook()

newSheet = newWorkbook['Sheet']

linhaAleatoria = random.randint(1, 1048576)

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    while newSheet.cell(row=linhaAleatoria, column=1).value != None:
        linhaAleatoria = random.randint(1, 1048576)
    
    if newSheet.cell(row=linhaAleatoria, column=1).value == None:
        newSheet.cell(row=linhaAleatoria, column=1).value = produceName
        print(newSheet.cell(row=linhaAleatoria, column=1))

newWorkbook.save('randomProduceSales.xlsx')

'''
letras = ''

letrasAleatoria = string.ascii_uppercase

qtdLetras = random.randint(1, newSheet.max_row)

print(qtdLetras)

for qtd in range(qtdLetras):
    letra = ''.join(random.choice(letrasAleatoria))
    letras += letra

print(letras)

#Este codigo gera letras maisuculas aleatorias de 1 ate 3 letras.
'''