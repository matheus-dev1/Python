#! python3

#atualizarPlanilha.py

#O programa deve atualizar valores de certos valores produtos.

#TODO:

#CTRL + / - Comentar tudo que esta selecionado

import openpyxl, os

os.chdir('C:\Matheus\Study\IT\Python\Arquivos Python\Criando e escrevendo em planilhas\Atualizando Valores de uma Planilha')

workbook = openpyxl.load_workbook('produceSales.xlsx')

sheet = workbook['Sheet']

atualizarProduto = {'Garlic': 3.07,
'Celery': 1.19,
'Lemon': 1.27}

for rowNum in range(2, sheet.max_row): # pula a primeira linha
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in atualizarProduto:
        sheet.cell(row=rowNum, column=2).value = atualizarProduto[produceName]
workbook.save('updatedProduceSales.xlsx')