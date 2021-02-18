#! python3

#MesclarCelulas.py

#O programa deve mesclar celulas.

#TODO:

import openpyxl, os

os.chdir('C:/Matheus/Study/IT/Python/Arquivos Python/Mescar Celulas e Separar Celulas/')

workbook = openpyxl.Workbook()

sheet = workbook['Sheet']

sheet['A1'] = 'Mesclar a celula A1 ate a D5'

sheet.merge_cells('A1:D5')
#Aqui eu vou mestar do A1 ate D5 ou seja, ele vai pegar B, C ate D

sheet['A6'] = 'mesclar a celula A6 ate a D10'

sheet.merge_cells('A6:D10')
#AQui eu vou mesclar da celula A6 ate a D10.

workbook.save('MesclarPlanilhas.xlsx')