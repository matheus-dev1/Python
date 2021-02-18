#! python3

#SeparaCelulasMesclada.

#O programa deve separar celulas mescladas.

#TODO:

import openpyxl, os

os.chdir('C:/Matheus/Study/IT/Python/Arquivos Python/Mescar Celulas e Separar Celulas/')

workbook = openpyxl.load_workbook('MesclarPlanilhas.xlsx')

sheet = workbook['Sheet']

sheet.unmerge_cells('A1:D5')

sheet.unmerge_cells('A6:D10')

workbook.save('SepararCelulasMescladas.xlsx')