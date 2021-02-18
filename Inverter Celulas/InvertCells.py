#! python3

#InvertCells.py

#O programa deve inverter uma planilha com dados na vertical e criar uma nova planila com os dados
#na horizontal.

#TODO:

#OBS: Pra que este programa desse certo, eu nao podia pegar e alterar a mesma planilha, por que isso
#daria conflito nos titulo, entao pra resolver isso nos criamos um outro workbook que ira armazenar
#os valores organizados da planilha com os valoers na vertical.

import openpyxl, os 

os.chdir("C:/Matheus/Study/IT/Python/Arquivos Python/Inverter Celulas")

workbook = openpyxl.load_workbook('InvertCells.xlsx')
sheet = workbook['Sheet1']

newWorkbook = openpyxl.Workbook()
newSheet = newWorkbook['Sheet']

for row in range(1, sheet.max_row + 1):
  for column in range(1, sheet.max_column + 1):
    newSheet.cell(column, row).value = sheet.cell(row, column).value

newWorkbook.save('InvertCellsOK.xlsx')