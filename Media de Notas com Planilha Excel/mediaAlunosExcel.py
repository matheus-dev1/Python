#! python3

#mediaAlunosExcel.py

#O programa deve recer os dados das notas de alunos e fazer suas medias.

#TODO: 1. Abrir o Workbook
#TODO: 2. Abrir a planilha.
#TODO: 3. Capturar os dados em forma de lista, sendo cada lista um aluno.
#TODO: 4. Fazer a media de cada aluno e exibir.
#TODO: 5. Fazer a media da turma.

import openpyxl

print("Opening workbook: notasAlunos.xlsx")

workbook = openpyxl.load_workbook(r"C:\Matheus\Study\IT\Python\Arquivos Python\Media de Notas com Planilha Excel\notasAlunos.xlsx")

sheet = workbook.get_sheet_by_name("Notas")

print('Reading Columns and Rows - Sheet: Notas...')

dadosTotal = []

for row in range(1, sheet.max_row + 1):
    notaAluno = []
    for column in range(1, sheet.max_column + 1):
        notaAluno.append(sheet.cell(row = row, column = column).value)
    dadosTotal.append(notaAluno)

medias = []

for aluno in range(len(dadosTotal)):
    notasTotal = 0
    for notas in dadosTotal[aluno]:
        notasTotal += notas
    media = notasTotal / len(dadosTotal[aluno])
    medias.append(media)

for aluno in range(len(medias)):
    print("Aluno: %s - Nota: %s" %(aluno + 1, medias[aluno]))