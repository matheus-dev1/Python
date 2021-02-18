#! python3

#PlanilhaParaTexto.py

#O programa deve pegar os dados de uma plnilha e colocar em um arquivo de texto.

#TODO:

import openpyxl, os

os.chdir('C:/Matheus/Study/IT/Python/Arquivos Python/Dados Planilha em Arquivo de Texto/')

workbook = openpyxl.load_workbook('TextoemPlanilha.xlsx')

sheet = workbook['Sheet']

TextFile = open('TextFile.txt', 'w')

for row in range(1, sheet.max_row + 1):
    TextFile.write(sheet.cell(row = row, column = 1).value)

TextFile.close()