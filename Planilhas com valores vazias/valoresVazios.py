#! python3

#valoresVazios.py

#O programa deve abrir um Workbook e uma planilha e verificar seus dados e se possui daos em branco ou invalidos
#do que nos queremos buscar exibir uma mensagem de alerta.

#TODO: 1. Abrir um workbook.
#TODO: 2. Abrir um planilha.
#TODO: 3. Capturar seus dados.
#TODO: 4. Armazenar em uma lista
#TODO: 5. Exibir cada item da lista e verificar se possui um valor vazio, caso possua, exibir uma mensagem.

import openpyxl

print("Opening workbook: valoresVazio.xlsx")

workbook = openpyxl.load_workbook(r'C:\Matheus\Study\IT\Python\Arquivos Python\Planilhas com valores vazias\valoresVazios.xlsx')

sheet = workbook.get_sheet_by_name("Dados")

print('Reading Columns and Rows - Sheet: Dados...')

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(row = row, column = column).value == None:
            print(str(sheet.cell(row = row, column = column).coordinate) + str(": Vazia"))
        elif type(sheet.cell(row = row, column = column).value) != int:
             print(str(sheet.cell(row = row, column = column).coordinate) + str(": Nao e um numero"))
        else:
            print(str(sheet.cell(row = row, column = column).coordinate) + ": " + str(sheet.cell(row = row, column = column).value))